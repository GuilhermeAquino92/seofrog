"""
seofrog/exporters/excel_exporter.py
Excel Exporter Enterprise do SEOFrog v0.2 - VERSÃO COMPATÍVEL
Funciona com ou sem módulos sheets especializados
"""

import pandas as pd
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

# Imports de dependências opcionais
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from seofrog.utils.logger import get_logger
from seofrog.core.exceptions import ExportException

class ExcelExporter:
    """
    Exportador Excel Enterprise compatível - funciona com ou sem módulos sheets
    """
    
    def __init__(self, output_dir: str = "seofrog_output"):
        self.output_dir = output_dir
        self.logger = get_logger('ExcelExporter')
        
        # Cria diretório se não existir
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        
        # Verifica dependências
        if not OPENPYXL_AVAILABLE:
            self.logger.warning("openpyxl não disponível. Install: pip install openpyxl")
    
    def export_results(self, crawl_data: List[Dict[str, Any]], filename: Optional[str] = None) -> str:
        """
        Exporta dados do crawl para Excel com múltiplas abas
        """
        
        if not crawl_data:
            self.logger.warning("Nenhum dado para exportar")
            return ""
        
        # Gera filename se não fornecido
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"seofrog_crawl_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            # Cria DataFrame principal
            df = pd.DataFrame(crawl_data)
            
            if df.empty:
                self.logger.error("DataFrame está vazio após criação")
                return ""
            
            # Define ordem das colunas
            column_order = self._get_column_order()
            available_columns = [col for col in column_order if col in df.columns]
            extra_columns = [col for col in df.columns if col not in column_order]
            final_columns = available_columns + sorted(extra_columns)
            
            # Reordena colunas e trata valores None
            df = df[final_columns].fillna('')
            
            # Determina engine
            engine = 'openpyxl' if OPENPYXL_AVAILABLE else 'xlsxwriter'
            
            # Cria arquivo Excel com múltiplas abas
            with pd.ExcelWriter(filepath, engine=engine) as writer:
                
                # === ABA 1: DADOS COMPLETOS ===
                df.to_excel(writer, sheet_name='Dados Completos', index=False)
                
                # === ABA 2: RESUMO EXECUTIVO ===
                self._create_summary_sheet(df, writer)
                
                # === ABAS DE PROBLEMAS ===
                self._create_status_problems_sheet(df, writer)
                self._create_title_problems_sheet(df, writer)
                self._create_meta_problems_sheet(df, writer)
                self._create_heading_problems_sheet(df, writer)
                self._create_h1_h2_missing_sheet(df, writer)
                self._create_image_problems_sheet(df, writer)
                self._create_technical_problems_sheet(df, writer)
                self._create_performance_problems_sheet(df, writer)
                
                # === ABA FINAL: ANÁLISE TÉCNICA ===
                self._create_technical_analysis_sheet(df, writer)
                
                # === FORMATAÇÃO ===
                if OPENPYXL_AVAILABLE:
                    self._format_workbook(writer)
            
            # Log estatísticas
            total_rows = len(df)
            total_columns = len(df.columns)
            file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
            
            self.logger.info(f"✅ Excel exportado: {filepath}")
            self.logger.info(f"📊 {total_rows:,} URLs × {total_columns} colunas ({file_size_mb:.1f} MB)")
            
            return filepath
            
        except Exception as e:
            error_msg = f"Erro exportando Excel: {e}"
            self.logger.error(error_msg)
            raise ExportException(error_msg, filename=filepath, format_type='xlsx')
    
    def _get_column_order(self) -> List[str]:
        """Define ordem das colunas por importância"""
        return [
            'url', 'final_url', 'status_code',
            'title', 'title_length', 'title_words',
            'meta_description', 'meta_description_length', 'meta_keywords',
            'h1_count', 'h1_text', 'h1_length',
            'h2_count', 'h3_count', 'h4_count', 'h5_count', 'h6_count',
            'internal_links_count', 'external_links_count', 'total_links_count',
            'images_count', 'images_without_alt', 'images_without_src',
            'word_count', 'character_count', 'text_ratio',
            'canonical_url', 'canonical_is_self', 'meta_robots',
            'has_viewport', 'has_charset', 'has_favicon',
            'schema_total_count', 'og_tags_count', 'twitter_tags_count',
            'response_time', 'content_length', 'content_type',
            'crawl_timestamp'
        ]
    
    def _safe_filter(self, df: pd.DataFrame, column: str, condition) -> pd.DataFrame:
        """Filtra DataFrame de forma segura"""
        if column not in df.columns:
            return pd.DataFrame()
        try:
            return df[condition].copy()
        except Exception as e:
            self.logger.warning(f"Erro filtrando coluna {column}: {e}")
            return pd.DataFrame()
    
    def _create_success_sheet(self, writer, sheet_name: str, message: str):
        """Cria aba de sucesso"""
        try:
            success_df = pd.DataFrame([[message]], columns=['Status'])
            success_df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            self.logger.warning(f"Erro criando aba {sheet_name}: {e}")
    
    def _create_error_sheet(self, writer, sheet_name: str, error_msg: str):
        """Cria aba de erro"""
        try:
            error_df = pd.DataFrame([[error_msg]], columns=['Erro'])
            error_df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            self.logger.warning(f"Erro criando aba de erro {sheet_name}: {e}")
    
    def _create_summary_sheet(self, df: pd.DataFrame, writer):
        """Cria aba de resumo executivo"""
        try:
            summary_data = []
            total_urls = len(df)
            
            if total_urls == 0:
                self._create_error_sheet(writer, 'Resumo Executivo', 'Nenhum dado para resumir')
                return
            
            summary_data.append(['Métrica', 'Valor', 'Percentual'])
            summary_data.append(['Total de URLs', total_urls, '100%'])
            
            # Status codes
            if 'status_code' in df.columns:
                try:
                    status_counts = df['status_code'].value_counts()
                    for status, count in status_counts.head(10).items():
                        percentage = f"{count/total_urls*100:.1f}%"
                        summary_data.append([f'Status {status}', count, percentage])
                except Exception as e:
                    self.logger.warning(f"Erro processando status codes: {e}")
            
            # Problemas SEO básicos
            if 'title' in df.columns:
                no_title = len(df[df['title'].fillna('') == ''])
                if no_title > 0:
                    summary_data.append(['URLs sem título', no_title, f"{no_title/total_urls*100:.1f}%"])
            
            if 'meta_description' in df.columns:
                no_meta = len(df[df['meta_description'].fillna('') == ''])
                if no_meta > 0:
                    summary_data.append(['URLs sem meta description', no_meta, f"{no_meta/total_urls*100:.1f}%"])
            
            if 'h1_count' in df.columns:
                no_h1 = len(df[df['h1_count'].fillna(0) == 0])
                if no_h1 > 0:
                    summary_data.append(['URLs sem H1', no_h1, f"{no_h1/total_urls*100:.1f}%"])
            
            # Cria DataFrame
            if len(summary_data) > 1:
                summary_df = pd.DataFrame(summary_data[1:], columns=summary_data[0])
                summary_df.to_excel(writer, sheet_name='Resumo Executivo', index=False)
            else:
                self._create_error_sheet(writer, 'Resumo Executivo', 'Erro na criação do resumo')
                
        except Exception as e:
            self.logger.error(f"Erro criando resumo: {e}")
            self._create_error_sheet(writer, 'Resumo Executivo', f'Erro: {str(e)}')
    
    def _create_status_problems_sheet(self, df: pd.DataFrame, writer):
        """Aba com problemas de status HTTP"""
        try:
            if 'status_code' not in df.columns:
                self._create_error_sheet(writer, 'Erros HTTP', 'Coluna status_code não encontrada')
                return
            
            status_problems = self._safe_filter(df, 'status_code', df['status_code'].fillna(0) != 200)
            
            if not status_problems.empty:
                status_problems = status_problems.sort_values('status_code')
                cols = ['url', 'status_code', 'final_url', 'response_time']
                available_cols = [col for col in cols if col in status_problems.columns]
                
                if available_cols:
                    status_problems[available_cols].to_excel(writer, sheet_name='Erros HTTP', index=False)
                else:
                    self._create_error_sheet(writer, 'Erros HTTP', 'Colunas necessárias não encontradas')
            else:
                self._create_success_sheet(writer, 'Erros HTTP', '✅ Nenhum erro HTTP encontrado!')
                
        except Exception as e:
            self.logger.error(f"Erro criando aba de erros HTTP: {e}")
            self._create_error_sheet(writer, 'Erros HTTP', f'Erro: {str(e)}')
    
    def _create_title_problems_sheet(self, df: pd.DataFrame, writer):
        """Aba com problemas de títulos"""
        try:
            title_issues = []
            
            # URLs sem título
            if 'title' in df.columns:
                no_title = self._safe_filter(df, 'title', df['title'].fillna('') == '')
                if not no_title.empty:
                    no_title['problema'] = 'Sem título'
                    title_issues.append(no_title)
            
            # Títulos muito longos
            if 'title_length' in df.columns:
                long_titles = self._safe_filter(df, 'title_length', df['title_length'].fillna(0) > 60)
                if not long_titles.empty:
                    long_titles['problema'] = 'Título muito longo (>60 chars)'
                    title_issues.append(long_titles)
            
            if title_issues:
                all_title_issues = pd.concat(title_issues, ignore_index=True)
                all_title_issues = all_title_issues.drop_duplicates(subset=['url'], keep='first')
                
                cols = ['url', 'problema', 'title', 'title_length']
                available_cols = [col for col in cols if col in all_title_issues.columns]
                
                if available_cols:
                    all_title_issues[available_cols].to_excel(writer, sheet_name='Problemas Títulos', index=False)
                else:
                    self._create_error_sheet(writer, 'Problemas Títulos', 'Colunas necessárias não encontradas')
            else:
                self._create_success_sheet(writer, 'Problemas Títulos', '✅ Nenhum problema de título!')
                
        except Exception as e:
            self.logger.error(f"Erro criando aba de problemas de título: {e}")
            self._create_error_sheet(writer, 'Problemas Títulos', f'Erro: {str(e)}')
    
    def _create_meta_problems_sheet(self, df: pd.DataFrame, writer):
        """Aba com problemas de meta description"""
        try:
            meta_issues = []
            
            if 'meta_description' in df.columns:
                no_meta = self._safe_filter(df, 'meta_description', df['meta_description'].fillna('') == '')
                if not no_meta.empty:
                    no_meta['problema'] = 'Sem meta description'
                    meta_issues.append(no_meta)
            
            if meta_issues:
                all_meta_issues = pd.concat(meta_issues, ignore_index=True)
                all_meta_issues = all_meta_issues.drop_duplicates(subset=['url'], keep='first')
                
                cols = ['url', 'problema', 'meta_description']
                available_cols = [col for col in cols if col in all_meta_issues.columns]
                
                if available_cols:
                    all_meta_issues[available_cols].to_excel(writer, sheet_name='Problemas Meta', index=False)
                else:
                    self._create_error_sheet(writer, 'Problemas Meta', 'Colunas necessárias não encontradas')
            else:
                self._create_success_sheet(writer, 'Problemas Meta', '✅ Nenhum problema de meta!')
                
        except Exception as e:
            self.logger.error(f"Erro criando aba de problemas de meta: {e}")
            self._create_error_sheet(writer, 'Problemas Meta', f'Erro: {str(e)}')
    
    def _create_heading_problems_sheet(self, df: pd.DataFrame, writer):
        """Aba com problemas de headings"""
        try:
            heading_issues = []
            
            if 'h1_count' in df.columns:
                no_h1 = self._safe_filter(df, 'h1_count', df['h1_count'].fillna(0) == 0)
                if not no_h1.empty:
                    no_h1['problema'] = 'Sem H1'
                    no_h1['criticidade'] = 'CRÍTICO'
                    heading_issues.append(no_h1)
            
            if heading_issues:
                all_heading_issues = pd.concat(heading_issues, ignore_index=True)
                all_heading_issues = all_heading_issues.drop_duplicates(subset=['url'], keep='first')
                
                cols = ['url', 'problema', 'criticidade', 'h1_count', 'h2_count']
                available_cols = [col for col in cols if col in all_heading_issues.columns]
                
                if available_cols:
                    all_heading_issues[available_cols].to_excel(writer, sheet_name='Problemas Headings', index=False)
                else:
                    self._create_error_sheet(writer, 'Problemas Headings', 'Colunas necessárias não encontradas')
            else:
                self._create_success_sheet(writer, 'Problemas Headings', '✅ Estrutura de headings adequada!')
                
        except Exception as e:
            self.logger.error(f"Erro criando aba de problemas de headings: {e}")
            self._create_error_sheet(writer, 'Problemas Headings', f'Erro: {str(e)}')
    
    def _create_h1_h2_missing_sheet(self, df: pd.DataFrame, writer):
        """Aba específica para URLs sem H1 e/ou H2"""
        try:
            h1_h2_issues = []
            
            # URLs sem H1
            if 'h1_count' in df.columns:
                no_h1 = self._safe_filter(df, 'h1_count', df['h1_count'].fillna(0) == 0)
                if not no_h1.empty:
                    no_h1['problema_h1_h2'] = 'Sem H1'
                    no_h1['criticidade'] = 'CRÍTICO'
                    h1_h2_issues.append(no_h1)
            
            # URLs sem H2
            if 'h2_count' in df.columns:
                no_h2 = self._safe_filter(df, 'h2_count', df['h2_count'].fillna(0) == 0)
                if not no_h2.empty:
                    no_h2['problema_h1_h2'] = 'Sem H2'
                    no_h2['criticidade'] = 'ALTO'
                    h1_h2_issues.append(no_h2)
            
            if h1_h2_issues:
                all_h1_h2_issues = pd.concat(h1_h2_issues, ignore_index=True)
                all_h1_h2_issues = all_h1_h2_issues.drop_duplicates(subset=['url'], keep='first')
                
                cols = ['url', 'problema_h1_h2', 'criticidade', 'h1_count', 'h2_count']
                available_cols = [col for col in cols if col in all_h1_h2_issues.columns]
                
                if available_cols:
                    all_h1_h2_issues[available_cols].to_excel(writer, sheet_name='H1 H2 Ausentes', index=False)
                else:
                    self._create_error_sheet(writer, 'H1 H2 Ausentes', 'Colunas necessárias não encontradas')
            else:
                self._create_success_sheet(writer, 'H1 H2 Ausentes', '✅ Estrutura de H1/H2 adequada!')
                
        except Exception as e:
            self.logger.error(f"Erro criando aba H1/H2 ausentes: {e}")
            self._create_error_sheet(writer, 'H1 H2 Ausentes', f'Erro: {str(e)}')
    
    def _create_image_problems_sheet(self, df: pd.DataFrame, writer):
        """Aba com problemas de imagens"""
        try:
            image_issues = []
            
            if 'images_without_alt' in df.columns:
                images_no_alt = self._safe_filter(df, 'images_without_alt', df['images_without_alt'].fillna(0) > 0)
                if not images_no_alt.empty:
                    images_no_alt['problema'] = 'Imagens sem ALT'
                    image_issues.append(images_no_alt)
            
            if image_issues:
                all_image_issues = pd.concat(image_issues, ignore_index=True)
                all_image_issues = all_image_issues.drop_duplicates(subset=['url'], keep='first')
                
                cols = ['url', 'problema', 'images_count', 'images_without_alt']
                available_cols = [col for col in cols if col in all_image_issues.columns]
                
                if available_cols:
                    all_image_issues[available_cols].to_excel(writer, sheet_name='Problemas Imagens', index=False)
                else:
                    self._create_error_sheet(writer, 'Problemas Imagens', 'Colunas necessárias não encontradas')
            else:
                self._create_success_sheet(writer, 'Problemas Imagens', '✅ Nenhum problema de imagem!')
                
        except Exception as e:
            self.logger.error(f"Erro criando aba de problemas de imagens: {e}")
            self._create_error_sheet(writer, 'Problemas Imagens', f'Erro: {str(e)}')
    
    def _create_technical_problems_sheet(self, df: pd.DataFrame, writer):
        """Aba com problemas técnicos"""
        try:
            technical_issues = []
            
            if 'canonical_url' in df.columns:
                no_canonical = self._safe_filter(df, 'canonical_url', df['canonical_url'].fillna('') == '')
                if not no_canonical.empty:
                    no_canonical['problema'] = 'Sem canonical'
                    technical_issues.append(no_canonical)
            
            if technical_issues:
                all_tech_issues = pd.concat(technical_issues, ignore_index=True)
                all_tech_issues = all_tech_issues.drop_duplicates(subset=['url'], keep='first')
                
                cols = ['url', 'problema', 'canonical_url']
                available_cols = [col for col in cols if col in all_tech_issues.columns]
                
                if available_cols:
                    all_tech_issues[available_cols].to_excel(writer, sheet_name='Problemas Técnicos', index=False)
                else:
                    self._create_error_sheet(writer, 'Problemas Técnicos', 'Colunas necessárias não encontradas')
            else:
                self._create_success_sheet(writer, 'Problemas Técnicos', '✅ Nenhum problema técnico!')
                
        except Exception as e:
            self.logger.error(f"Erro criando aba de problemas técnicos: {e}")
            self._create_error_sheet(writer, 'Problemas Técnicos', f'Erro: {str(e)}')
    
    def _create_performance_problems_sheet(self, df: pd.DataFrame, writer):
        """Aba com problemas de performance"""
        try:
            perf_issues = []
            
            if 'response_time' in df.columns:
                slow_pages = self._safe_filter(df, 'response_time', df['response_time'].fillna(0) > 3)
                if not slow_pages.empty:
                    slow_pages['problema'] = 'Página lenta (>3s)'
                    perf_issues.append(slow_pages)
            
            if perf_issues:
                all_perf_issues = pd.concat(perf_issues, ignore_index=True)
                all_perf_issues = all_perf_issues.drop_duplicates(subset=['url'], keep='first')
                
                # Ordena por tempo de resposta
                if 'response_time' in all_perf_issues.columns:
                    all_perf_issues = all_perf_issues.sort_values('response_time', ascending=False)
                
                cols = ['url', 'problema', 'response_time']
                available_cols = [col for col in cols if col in all_perf_issues.columns]
                
                if available_cols:
                    all_perf_issues[available_cols].to_excel(writer, sheet_name='Problemas Performance', index=False)
                else:
                    self._create_error_sheet(writer, 'Problemas Performance', 'Colunas necessárias não encontradas')
            else:
                self._create_success_sheet(writer, 'Problemas Performance', '✅ Nenhum problema de performance!')
                
        except Exception as e:
            self.logger.error(f"Erro criando aba de problemas de performance: {e}")
            self._create_error_sheet(writer, 'Problemas Performance', f'Erro: {str(e)}')
    
    def _create_technical_analysis_sheet(self, df: pd.DataFrame, writer):
        """Cria aba com análise técnica"""
        try:
            tech_columns = ['url', 'status_code', 'canonical_url', 'meta_robots', 'response_time']
            available_tech_cols = [col for col in tech_columns if col in df.columns]
            
            if available_tech_cols:
                tech_df = df[available_tech_cols].copy()
                
                # Adiciona análise técnica resumida
                tech_df['analise_tecnica'] = tech_df.apply(self._generate_technical_analysis, axis=1)
                
                tech_df.to_excel(writer, sheet_name='Análise Técnica', index=False)
            else:
                self._create_error_sheet(writer, 'Análise Técnica', 'Nenhuma coluna técnica encontrada')
                
        except Exception as e:
            self.logger.error(f"Erro criando aba técnica: {e}")
            self._create_error_sheet(writer, 'Análise Técnica', f'Erro: {str(e)}')
    
    def _generate_technical_analysis(self, row) -> str:
        """Gera análise técnica resumida para cada URL"""
        try:
            issues = []
            
            # Verifica problemas técnicos
            if row.get('status_code', 200) != 200:
                issues.append(f"Status {row.get('status_code')}")
            
            if not row.get('canonical_url', ''):
                issues.append("Sem canonical")
            
            if row.get('response_time', 0) > 3:
                issues.append("Lenta")
            
            meta_robots = row.get('meta_robots', '').lower()
            if 'noindex' in meta_robots:
                issues.append("NOINDEX")
            
            # Retorna resumo
            if issues:
                return '; '.join(issues)
            else:
                return '✅ OK'
                
        except Exception:
            return 'Erro na análise'
    
    def _format_workbook(self, writer):
        """Aplica formatação ao workbook"""
        try:
            if not OPENPYXL_AVAILABLE:
                self.logger.warning("openpyxl não disponível para formatação")
                return
            
            workbook = writer.book
            
            # Define estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Aplica formatação em todas as abas
            for sheet_name in workbook.sheetnames:
                try:
                    worksheet = workbook[sheet_name]
                    
                    # Verifica se a aba tem dados
                    if worksheet.max_row < 1:
                        continue
                    
                    # Formata header (primeira linha)
                    for cell in worksheet[1]:
                        if cell.value:  # Só formata células com conteúdo
                            cell.font = header_font
                            cell.fill = header_fill
                    
                    # Congela primeira linha
                    worksheet.freeze_panes = "A2"
                    
                    # Auto-ajusta largura das colunas (limitado a 50 chars)
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        # Define largura (mínimo 10, máximo 50)
                        adjusted_width = min(max(max_length, 10), 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                        
                except Exception as e:
                    self.logger.warning(f"Erro formatando aba '{sheet_name}': {e}")
                    continue
            
            self.logger.info("✅ Formatação Excel aplicada com sucesso")
            
        except Exception as e:
            self.logger.warning(f"Erro geral na formatação: {e}")


# Função auxiliar para uso externo
def export_to_excel(crawl_data: List[Dict[str, Any]], 
                   output_dir: str = "seofrog_output", 
                   filename: Optional[str] = None) -> str:
    """
    Função auxiliar para exportar dados para Excel
    
    Args:
        crawl_data: Lista de dicionários com dados do crawl
        output_dir: Diretório de saída
        filename: Nome do arquivo (opcional)
    
    Returns:
        Caminho do arquivo gerado
    """
    exporter = ExcelExporter(output_dir)
    return exporter.export_results(crawl_data, filename)


if __name__ == "__main__":
    print("🐸 SEOFrog Excel Exporter v0.2 - Versão Compatível")
    print("✅ Funciona independente da estrutura modular")